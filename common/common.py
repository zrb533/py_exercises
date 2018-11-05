# -*- coding: utf-8 -*-#
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font

def logs(error,log_path):
    type_error = "错误类型：" + str(type(error))
    content_error = "错误内容：" + str(error)
    currency_time = "时间：" + datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    error = currency_time + ' ' * 5 + type_error + ' ' * 5 + content_error + ' ' * 5 + __file__
    file = open(log_path, 'a')
    file.write(error + '\n')
    file.close()

class parseExcel(object):


    def __init__(self,excelPath):
        self.excelPath = excelPath
        self.workbook = load_workbook(self.excelPath)
        self.sheet = self.workbook.active

    # 使用index来获取当前操作的sheet
    def set_sheet_by_index(self,sheet_index):
        sheet_name = self.workbook.get_sheet_names()[sheet_index]
        self.sheet = self.workbook.get_sheet_by_name(sheet_name)
        return self.sheet

    # 使用sheetname来获取当前操作的sheet
    def set_sheet_by_name(self,sheet_name):
        sheet = self.workbook[sheet_name]
        self.sheet = sheet
        return self.sheet

    # 获取当前sheet最大行数
    def get_max_row_no(self):
        return self.sheet.max_row

    # 获取当前sheet最大行数
    def get_max_column_no(self):
        return self.sheet.max_column

    # 通过行号和列号获取指定的单元格的值
    def get_cell_content(self, row_no, col_no):
        return self.sheet.cell(row=row_no, column=col_no).value
